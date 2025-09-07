from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import tempfile
import os

from src.models.user import db
from src.models.indicador import Indicador
from src.models.indicacao import Indicacao, StatusRecompensa

relatorios_bp = Blueprint('relatorios', __name__)

def format_currency(value_in_cents):
    """Converte centavos para formato de moeda brasileira"""
    if value_in_cents is None:
        return "R$ 0,00"
    return f"R$ {value_in_cents / 100:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def format_phone(phone):
    """Formata telefone para exibição"""
    if not phone:
        return ""
    # Remove +55 e formata como (XX) XXXXX-XXXX
    clean_phone = phone.replace('+55', '')
    if len(clean_phone) == 11:
        return f"({clean_phone[:2]}) {clean_phone[2:7]}-{clean_phone[7:]}"
    return phone

@relatorios_bp.route('/dashboard-stats', methods=['GET'])
def get_dashboard_stats():
    """Retorna estatísticas para o dashboard com filtros opcionais"""
    try:
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        indicador_id = request.args.get('indicador_id')
        
        # Query base
        query = db.session.query(Indicacao)
        
        # Aplicar filtros
        if data_inicio:
            query = query.filter(Indicacao.data_indicacao >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
        if data_fim:
            query = query.filter(Indicacao.data_indicacao <= datetime.strptime(data_fim, '%Y-%m-%d').date())
        if indicador_id:
            query = query.filter(Indicacao.indicador_id == indicador_id)
        
        # Calcular estatísticas
        total_indicacoes = query.count()
        total_vendas = query.filter(Indicacao.gerou_venda == True).count()
        faturamento_total = query.filter(Indicacao.gerou_venda == True).with_entities(
            func.sum(Indicacao.faturamento_gerado)
        ).scalar() or 0
        
        # Total de indicadores únicos
        total_indicadores = db.session.query(Indicador).count()
        if indicador_id:
            total_indicadores = 1
        
        # Taxa de conversão
        taxa_conversao = (total_vendas / total_indicacoes * 100) if total_indicacoes > 0 else 0
        
        return jsonify({
            'total_indicacoes': total_indicacoes,
            'total_indicadores': total_indicadores,
            'total_vendas': total_vendas,
            'taxa_conversao': round(taxa_conversao, 1),
            'faturamento_total': faturamento_total
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/performance-indicadores', methods=['GET'])
def get_performance_indicadores():
    """Retorna performance detalhada por indicador"""
    try:
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # Query com join para pegar dados do indicador e suas indicações
        query = db.session.query(
            Indicador.id,
            Indicador.nome,
            Indicador.empresa,
            Indicador.telefone,
            Indicador.email,
            func.count(Indicacao.id).label('total_indicacoes'),
            func.sum(func.case((Indicacao.gerou_venda == True, 1), else_=0)).label('total_vendas'),
            func.sum(func.case((Indicacao.gerou_venda == True, Indicacao.faturamento_gerado), else_=0)).label('faturamento_total')
        ).outerjoin(Indicacao).group_by(Indicador.id)
        
        # Aplicar filtros de data nas indicações
        if data_inicio or data_fim:
            if data_inicio:
                query = query.filter(or_(
                    Indicacao.data_indicacao >= datetime.strptime(data_inicio, '%Y-%m-%d').date(),
                    Indicacao.id.is_(None)
                ))
            if data_fim:
                query = query.filter(or_(
                    Indicacao.data_indicacao <= datetime.strptime(data_fim, '%Y-%m-%d').date(),
                    Indicacao.id.is_(None)
                ))
        
        resultados = query.all()
        
        performance = []
        for resultado in resultados:
            total_indicacoes = resultado.total_indicacoes or 0
            total_vendas = resultado.total_vendas or 0
            faturamento_total = resultado.faturamento_total or 0
            taxa_conversao = (total_vendas / total_indicacoes * 100) if total_indicacoes > 0 else 0
            
            performance.append({
                'id': str(resultado.id),
                'nome': resultado.nome,
                'empresa': resultado.empresa,
                'telefone': format_phone(resultado.telefone),
                'email': resultado.email,
                'total_indicacoes': total_indicacoes,
                'total_vendas': total_vendas,
                'taxa_conversao': round(taxa_conversao, 1),
                'faturamento_total': faturamento_total
            })
        
        # Ordenar por faturamento total (decrescente)
        performance.sort(key=lambda x: x['faturamento_total'], reverse=True)
        
        return jsonify(performance)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/export/excel', methods=['GET'])
def export_excel():
    """Exporta dados para Excel com múltiplas abas"""
    try:
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        indicador_id = request.args.get('indicador_id')
        tipo_relatorio = request.args.get('tipo', 'completo')  # completo, indicacoes, indicadores
        
        # Criar workbook
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if tipo_relatorio in ['completo', 'indicacoes']:
            # Aba 1: Indicações
            ws_indicacoes = wb.active
            ws_indicacoes.title = "Indicações"
            
            # Query das indicações
            query = db.session.query(
                Indicacao.data_indicacao,
                Indicador.nome.label('indicador_nome'),
                Indicador.empresa.label('indicador_empresa'),
                Indicador.telefone.label('indicador_telefone'),
                Indicacao.nome_indicado,
                Indicacao.telefone_indicado,
                Indicacao.gerou_venda,
                Indicacao.faturamento_gerado,
                Indicacao.status_recompensa,
                Indicacao.observacoes
            ).join(Indicador)
            
            # Aplicar filtros
            if data_inicio:
                query = query.filter(Indicacao.data_indicacao >= datetime.strptime(data_inicio, '%Y-%m-%d').date())
            if data_fim:
                query = query.filter(Indicacao.data_indicacao <= datetime.strptime(data_fim, '%Y-%m-%d').date())
            if indicador_id:
                query = query.filter(Indicacao.indicador_id == indicador_id)
            
            indicacoes = query.order_by(Indicacao.data_indicacao.desc()).all()
            
            # Headers
            headers = [
                'Data', 'Indicador', 'Empresa', 'Tel. Indicador', 'Nome Indicado', 
                'Tel. Indicado', 'Gerou Venda', 'Faturamento', 'Status Recompensa', 'Observações'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_indicacoes.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Dados
            for row, indicacao in enumerate(indicacoes, 2):
                ws_indicacoes.cell(row=row, column=1, value=indicacao.data_indicacao.strftime('%d/%m/%Y'))
                ws_indicacoes.cell(row=row, column=2, value=indicacao.indicador_nome)
                ws_indicacoes.cell(row=row, column=3, value=indicacao.indicador_empresa or '')
                ws_indicacoes.cell(row=row, column=4, value=format_phone(indicacao.indicador_telefone))
                ws_indicacoes.cell(row=row, column=5, value=indicacao.nome_indicado)
                ws_indicacoes.cell(row=row, column=6, value=format_phone(indicacao.telefone_indicado))
                ws_indicacoes.cell(row=row, column=7, value='Sim' if indicacao.gerou_venda else 'Não')
                ws_indicacoes.cell(row=row, column=8, value=format_currency(indicacao.faturamento_gerado))
                ws_indicacoes.cell(row=row, column=9, value=indicacao.status_recompensa.value)
                ws_indicacoes.cell(row=row, column=10, value=indicacao.observacoes or '')
                
                # Aplicar bordas
                for col in range(1, 11):
                    ws_indicacoes.cell(row=row, column=col).border = border
            
            # Ajustar largura das colunas
            column_widths = [12, 25, 20, 15, 25, 15, 12, 15, 18, 30]
            for col, width in enumerate(column_widths, 1):
                ws_indicacoes.column_dimensions[get_column_letter(col)].width = width
        
        if tipo_relatorio in ['completo', 'indicadores']:
            # Aba 2: Performance por Indicador
            if tipo_relatorio == 'completo':
                ws_performance = wb.create_sheet(title="Performance Indicadores")
            else:
                ws_performance = wb.active
                ws_performance.title = "Performance Indicadores"
            
            # Query de performance
            query = db.session.query(
                Indicador.nome,
                Indicador.empresa,
                Indicador.telefone,
                Indicador.email,
                func.count(Indicacao.id).label('total_indicacoes'),
                func.sum(func.case((Indicacao.gerou_venda == True, 1), else_=0)).label('total_vendas'),
                func.sum(func.case((Indicacao.gerou_venda == True, Indicacao.faturamento_gerado), else_=0)).label('faturamento_total')
            ).outerjoin(Indicacao).group_by(Indicador.id)
            
            # Aplicar filtros de data
            if data_inicio or data_fim:
                if data_inicio:
                    query = query.filter(or_(
                        Indicacao.data_indicacao >= datetime.strptime(data_inicio, '%Y-%m-%d').date(),
                        Indicacao.id.is_(None)
                    ))
                if data_fim:
                    query = query.filter(or_(
                        Indicacao.data_indicacao <= datetime.strptime(data_fim, '%Y-%m-%d').date(),
                        Indicacao.id.is_(None)
                    ))
            
            performance = query.order_by(func.sum(func.case((Indicacao.gerou_venda == True, Indicacao.faturamento_gerado), else_=0)).desc()).all()
            
            # Headers
            headers = [
                'Nome', 'Empresa', 'Telefone', 'Email', 'Total Indicações', 
                'Total Vendas', 'Taxa Conversão (%)', 'Faturamento Total'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_performance.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Dados
            for row, indicador in enumerate(performance, 2):
                total_indicacoes = indicador.total_indicacoes or 0
                total_vendas = indicador.total_vendas or 0
                faturamento_total = indicador.faturamento_total or 0
                taxa_conversao = (total_vendas / total_indicacoes * 100) if total_indicacoes > 0 else 0
                
                ws_performance.cell(row=row, column=1, value=indicador.nome)
                ws_performance.cell(row=row, column=2, value=indicador.empresa or '')
                ws_performance.cell(row=row, column=3, value=format_phone(indicador.telefone))
                ws_performance.cell(row=row, column=4, value=indicador.email or '')
                ws_performance.cell(row=row, column=5, value=total_indicacoes)
                ws_performance.cell(row=row, column=6, value=total_vendas)
                ws_performance.cell(row=row, column=7, value=f"{taxa_conversao:.1f}%")
                ws_performance.cell(row=row, column=8, value=format_currency(faturamento_total))
                
                # Aplicar bordas
                for col in range(1, 9):
                    ws_performance.cell(row=row, column=col).border = border
            
            # Ajustar largura das colunas
            column_widths = [25, 20, 15, 25, 15, 12, 15, 18]
            for col, width in enumerate(column_widths, 1):
                ws_performance.column_dimensions[get_column_letter(col)].width = width
        
        # Salvar em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Nome do arquivo baseado no tipo e filtros
        filename = f"relatorio_{tipo_relatorio}"
        if data_inicio and data_fim:
            filename += f"_{data_inicio}_a_{data_fim}"
        elif data_inicio:
            filename += f"_desde_{data_inicio}"
        elif data_fim:
            filename += f"_ate_{data_fim}"
        filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        # Limpar arquivo temporário
        try:
            if 'temp_file' in locals():
                os.unlink(temp_file.name)
        except:
            pass

